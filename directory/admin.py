# from django.contrib import admin
# from django.utils.html import format_html
# from django.http import HttpResponse
# from openpyxl import Workbook
# from .models import Member, FamilyMember, Area

# # =========================================================================
# # 1. EXCEL EXPORT ACTIONS
# # =========================================================================

# # A. MAIN MEMBER + INLINE FAMILY MEMBERS EXPORT
# @admin.action(description='Selected Parivaro ka poora data Excel mein export karein')
# def export_member_directory_to_excel(modeladmin, request, queryset):
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename="Samaj_Parivar_Directory_Full.xlsx"'
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Parivar Directory"
    
#     headers = [
#         'Registration No', 'Category / Relation', 'Name', ''
#         'Phone Number', 
#         'Date of Birth', 'Age', 'Marital Status', 'Spouse Name', 
#         'Detailed Address', 'Area', 'Gotra'
#     ]
#     ws.append(headers)
    
#     for head_member in queryset:
#         reg_no = getattr(head_member, 'registration_no', '')
#         common_address = getattr(head_member, 'detailed_address', '')
        
#         area_obj = getattr(head_member, 'area', '')
#         common_area = str(area_obj) if area_obj else ''
#         common_gotra = getattr(head_member, 'gotra', '')
        
#         # Mukhiya Row (Calculates from 'age' property)
#         head_dob = getattr(head_member, 'dob', None)
#         head_age = f"{head_member.age} Yrs" if head_member.age else "--"
        
#         ws.append([
#             reg_no, 'Main Member (Head)', getattr(head_member, 'name', ''),
#             getattr(head_member, 'phone_number', ''),
#             head_dob.strftime('%Y-%m-%d') if head_dob else '', head_age,
#             getattr(head_member, 'marital_status', ''), getattr(head_member, 'spouse_name', ''),
#             common_address, common_area, common_gotra
#         ])
        
#         # Family Members Rows (Using your exact related_name 'family_members')
#         family_members = head_member.family_members.all() if hasattr(head_member, 'family_members') else []
#         for f_member in family_members:
#             f_dob = getattr(f_member, 'dob', None)
#             f_age = f"{f_member.age} Yrs" if f_member.age else "--"
            
#             ws.append([
#                 reg_no, 'Family Member', getattr(f_member, 'name', ''),
#                 getattr(f_member, 'phone', ''),
#                 f_dob.strftime('%Y-%m-%d') if f_dob else '', f_age,
#                 getattr(f_member, 'marital_status', ''), getattr(f_member, 'spouse_name', ''),
#                 common_address, common_area, common_gotra
#             ])
            
#     wb.save(response)
#     return response


# # B. ONLY FAMILY MEMBERS EXPORT
# @admin.action(description='Selected Family Members ko Excel mein export karein')
# def export_only_family_members_to_excel(modeladmin, request, queryset):
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename="Family_Members_Only.xlsx"'
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Family Members"
    
#     headers = ['Name', 'Phone Number', 'Date of Birth', 'Age', 'Marital Status', 'Spouse Name', 'Address', 'Area', 'Gotra']
#     ws.append(headers)
    
#     for f_member in queryset:
#         main_family = getattr(f_member, 'member', None)
#         f_dob = getattr(f_member, 'dob', None)
#         f_age = f"{f_member.age} Yrs" if f_member.age else "--"
        
#         common_address = getattr(main_family, 'detailed_address', '') if main_family else ''
#         area_obj = getattr(main_family, 'area', '') if main_family else ''
#         common_area = str(area_obj) if area_obj else ''
#         common_gotra = getattr(main_family, 'gotra', '') if main_family else ''
        
#         ws.append([
#             getattr(f_member, 'name', ''), getattr(f_member, 'phone', ''),
#             f_dob.strftime('%Y-%m-%d') if f_dob else '', f_age,
#             getattr(f_member, 'marital_status', ''), getattr(f_member, 'spouse_name', ''),
#             common_address, common_area, common_gotra
#         ])
        
#     wb.save(response)
#     return response


# # =========================================================================
# # 2. INLINES AND ADMIN CLASSES SETUP
# # =========================================================================

# # Family Member Inline
# class FamilyMemberInline(admin.TabularInline if hasattr(admin, 'TabularInline') else admin.TabularInline):
#     model = FamilyMember
#     extra = 0
    
#     # 🌟 readonly_fields mein humne display_age aur display_spouse_age ke sath 'dob' aur 'spouse_dob' ko bhi safe side rakh diya hai
#     readonly_fields = ['display_age', 'display_spouse_age']
    
#     # 🌟 Fields ka sequence ekdum accurate set kiya hai taaki admin panel ko dikhne mein koi dikkat na ho
#     fields = [
#         'name', 'phone', 'dob', 'display_age', 'occupation', 
#         'marital_status', 'spouse_name', 'spouse_dob', 'display_spouse_age'
#     ]

#     def display_age(self, obj):
#         # 🌟 Agar obj hai aur uski dob database mein hai, toh direct calculation fallback lagaya hai
#         if obj and obj.dob:
#             try:
#                 from datetime import date
#                 today = date.today()
#                 calculated_age = today.year - obj.dob.year - ((today.month, today.day) < (obj.dob.month, obj.dob.day))
#                 return f"{calculated_age} Yrs"
#             except Exception as e:
#                 return "--"
#         return "--"
#     display_age.short_description = 'Member Age'

#     def display_spouse_age(self, obj):
#         # 🌟 Spouse ke liye bhi direct fail-safe calculation logic
#         if obj and obj.spouse_dob:
#             try:
#                 from datetime import date
#                 today = date.today()
#                 calculated_spouse_age = today.year - obj.spouse_dob.year - ((today.month, today.day) < (obj.spouse_dob.month, obj.spouse_dob.day))
#                 return f"{calculated_spouse_age} Yrs"
#             except Exception as e:
#                 return "--"
#         return "--"
#     display_spouse_age.short_description = 'Spouse Age'
# # Main Member Admin (Parivar Head ke liye)
# @admin.register(Member)
# class MemberAdmin(admin.ModelAdmin):
#     # list_display ko ekdum safe rakh rahe hain, bina kisi complex HTML loop ke
#     list_display = [
#         'registration_no', 'name', 'phone_number', 'pin', 
#         'get_head_age', 'spouse_name', 'get_head_spouse_age', 
#         'get_family_members', 'area', 'gotra', 'marital_status'
#     ]
#     ordering = ['registration_no']
#     readonly_fields = ['registration_no', 'get_head_age', 'get_head_spouse_age']
    
#     fieldsets = (
#         ('Basic Information', {
#             'fields': ('name', 'phone_number', 'dob', 'get_head_age', 'marital_status', 'pin', 'is_head')
#         }),
#         ('Spouse Details', {
#             'fields': ('spouse_name', 'spouse_phone', 'spouse_dob', 'get_head_spouse_age')
#         }),
#         ('Address & Location', {
#             'fields': ('gotra', 'area', 'detailed_address')
#         }),
#     )
    
#     inlines = [FamilyMemberInline]
#     actions = [export_member_directory_to_excel]  

#     # 1. Head Age
#     def get_head_age(self, obj):
#         return f"{obj.age} Yrs" if obj.age else "--"
#     get_head_age.short_description = 'Age'

#     # 2. Spouse Age
#     def get_head_spouse_age(self, obj):
#         if obj.spouse_dob:
#             from datetime import date
#             today = date.today()
#             calculated_spouse_age = today.year - obj.spouse_dob.year - ((today.month, today.day) < (obj.spouse_dob.month, obj.spouse_dob.day))
#             return f"{calculated_spouse_age} Yrs"
#         return "--"
#     get_head_spouse_age.short_description = 'Spouse Age'

#     # 🎯 3. YEH EKDOM SAFE TARIKA HAI: Saare family members ke naam comma se print honge, bina crash kiye
#     def get_family_members(self, obj):
#         members = obj.family_members.all()
#         if members:
#             return ", ".join([m.name for m in members])
#         return "--"
#     get_family_members.short_description = 'Family Members'
# @admin.register(FamilyMember)
# class FamilyMemberAdmin(admin.ModelAdmin):
#     list_display = ['name', 'phone', 'marital_status', 'spouse_name', 'member']  
#     actions = [export_only_family_members_to_excel]


# # Area Registration
# admin.site.register(Area)

from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Member, FamilyMember, Area

# =========================================================================
# 1. EXCEL EXPORT ACTIONS
# =========================================================================

# A. MAIN MEMBER + INLINE FAMILY MEMBERS EXPORT
@admin.action(description='Selected Parivaro ka poora data Excel mein export karein')
def export_member_directory_to_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Samaj_Parivar_Directory_Full.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Parivar Directory"
    
    # 🌟 FIXED: Added missing commas and aligned headers perfectly with columns
    headers = [
        'Registration No', 'Category / Relation', 'Name', 'Phone Number', 
        'Date of Birth', 'Occupation', 'Age', 'Marital Status', 'Spouse Name', 
        'Detailed Address', 'Area', 'Gotra'
    ]
    ws.append(headers)
    
    for head_member in queryset:
        reg_no = getattr(head_member, 'registration_no', '')
        common_address = getattr(head_member, 'detailed_address', '')
        
        area_obj = getattr(head_member, 'area', '')
        common_area = str(area_obj) if area_obj else ''
        common_gotra = getattr(head_member, 'gotra', '')
        
        head_dob = getattr(head_member, 'dob', None)
        head_age = f"{head_member.age} Yrs" if head_member.age else "--"
        
        ws.append([
            reg_no, 'Main Member (Head)', getattr(head_member, 'name', ''),
            getattr(head_member, 'phone_number', ''),
            head_dob.strftime('%Y-%m-%d') if head_dob else '', 
            'Head / Business',  # Mukhiya occupation fallback
            head_age,
            getattr(head_member, 'marital_status', ''), getattr(head_member, 'spouse_name', ''),
            common_address, common_area, common_gotra
        ])
        
        family_members = head_member.family_members.all() if hasattr(head_member, 'family_members') else []
        for f_member in family_members:
            f_dob = getattr(f_member, 'dob', None)
            f_age = f"{f_member.age} Yrs" if f_member.age else "--"
            
            ws.append([
                reg_no, 'Family Member', getattr(f_member, 'name', ''),
                getattr(f_member, 'phone', ''),
                f_dob.strftime('%Y-%m-%d') if f_dob else '', 
                getattr(f_member, 'occupation', '--'),  # Excel export mein bhi occupation jodh diya
                f_age,
                getattr(f_member, 'marital_status', ''), getattr(f_member, 'spouse_name', ''),
                common_address, common_area, common_gotra
            ])
            
    wb.save(response)
    return response


# B. ONLY FAMILY MEMBERS EXPORT
@admin.action(description='Selected Family Members ko Excel mein export karein')
def export_only_family_members_to_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Family_Members_Only.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Family Members"
    
    headers = ['Name', 'Phone Number', 'Date of Birth', 'Occupation', 'Age', 'Marital Status', 'Spouse Name', 'Address', 'Area', 'Gotra']
    ws.append(headers)
    
    for f_member in queryset:
        main_family = getattr(f_member, 'member', None)
        f_dob = getattr(f_member, 'dob', None)
        f_age = f"{f_member.age} Yrs" if f_member.age else "--"
        
        common_address = getattr(main_family, 'detailed_address', '') if main_family else ''
        area_obj = getattr(main_family, 'area', '') if main_family else ''
        common_area = str(area_obj) if area_obj else ''
        common_gotra = getattr(main_family, 'gotra', '') if main_family else ''
        
        ws.append([
            getattr(f_member, 'name', ''), getattr(f_member, 'phone', ''),
            f_dob.strftime('%Y-%m-%d') if f_dob else '', 
            getattr(f_member, 'occupation', '--'), # Inline model handler
            f_age,
            getattr(f_member, 'marital_status', ''), getattr(f_member, 'spouse_name', ''),
            common_address, common_area, common_gotra
        ])
        
    wb.save(response)
    return response


# =========================================================================
# 2. INLINES AND ADMIN CLASSES SETUP
# =========================================================================

class FamilyMemberInline(admin.TabularInline):
    model = FamilyMember
    extra = 0
    readonly_fields = ['display_age', 'display_spouse_age']
    
    fields = [
        'name', 'phone', 'dob', 'display_age', 'occupation', 
        'marital_status', 'spouse_name', 'spouse_dob', 'display_spouse_age'
    ]

    def display_age(self, obj):
        if obj and obj.dob:
            try:
                from datetime import date
                today = date.today()
                calculated_age = today.year - obj.dob.year - ((today.month, today.day) < (obj.dob.month, obj.dob.day))
                return f"{calculated_age} Yrs"
            except:
                return "--"
        return "--"
    display_age.short_description = 'Member Age'

    def display_spouse_age(self, obj):
        if obj and obj.spouse_dob:
            try:
                from datetime import date
                today = date.today()
                calculated_spouse_age = today.year - obj.spouse_dob.year - ((today.month, today.day) < (obj.spouse_dob.month, obj.spouse_dob.day))
                return f"{calculated_spouse_age} Yrs"
            except:
                return "--"
        return "--"
    display_spouse_age.short_description = 'Spouse Age'


@admin.register(Member)
class MemberAdmin(admin.ModelAdmin):
    list_display = [
        'registration_no', 'name', 'phone_number','pin', 'head_occupation',
        'spouse_name', 'spouse_occupation',                           
        'get_head_age', 'get_family_members', 'area', 'gotra'
    ]
    ordering = ['registration_no']
    readonly_fields = ['registration_no', 'get_head_age', 'get_head_spouse_age']
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'phone_number', 'dob', 'head_occupation', 'get_head_age', 'marital_status', 'pin', 'is_head')
        }),
        ('Spouse Details', {
            'fields': ('spouse_name', 'spouse_phone', 'spouse_dob', 'spouse_occupation', 'get_head_spouse_age')
        }),
        ('Address & Location', {
            'fields': ('gotra', 'area', 'detailed_address')
        }),
    )
    
    inlines = [FamilyMemberInline]
    actions = [export_member_directory_to_excel]  

    def get_head_age(self, obj):
        return f"{obj.age} Yrs" if obj.age else "--"
    get_head_age.short_description = 'Age'

    def get_head_spouse_age(self, obj):
        if obj.spouse_dob:
            from datetime import date
            today = date.today()
            calculated_spouse_age = today.year - obj.spouse_dob.year - ((today.month, today.day) < (obj.spouse_dob.month, obj.spouse_dob.day))
            return f"{calculated_spouse_age} Yrs"
        return "--"
    get_head_spouse_age.short_description = 'Spouse Age'

    def get_family_members(self, obj):
        members = obj.family_members.all()
        if members:
            return ", ".join([m.name for m in members])
        return "--"
    get_family_members.short_description = 'Family Members'

@admin.register(FamilyMember)
class FamilyMemberAdmin(admin.ModelAdmin):
    # 🌟 Added occupation to standalone view as well for backup
    list_display = ['name', 'phone', 'dob', 'occupation', 'marital_status', 'spouse_name', 'member']  
    actions = [export_only_family_members_to_excel]


admin.site.register(Area)